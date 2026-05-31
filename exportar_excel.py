"""
exportar_excel.py — Exporta todos os dados do bolão para Excel.

Uso:
    python exportar_excel.py
    python exportar_excel.py --db meu_banco.db --out relatorio.xlsx

Gera 4 abas:
    1. Ranking          — classificação geral com pontuação detalhada
    2. Palpites Jogos   — todos os palpites de placar por participante
    3. Palpites Classif — palpites de classificação do grupo
    4. Configurações    — placares reais e classificação final registrados
"""

import sqlite3, os, sys, argparse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Cores ──────────────────────────────────────────────────
COR_ROXO_ESC  = "1A1464"
COR_ROXO_MED  = "2D1B69"
COR_VERDE     = "009C3B"
COR_AMARELO   = "FFDF00"
COR_BRANCO    = "FFFFFF"
COR_CINZA_CLR = "F2F2F2"
COR_CINZA_BD  = "CCCCCC"
COR_VERMELHO  = "C8102E"
COR_OURO      = "FFD700"
COR_PRATA     = "C0C0C0"
COR_BRONZE    = "CD7F32"
COR_VERDE_CLR = "E8F5E9"
COR_VERM_CLR  = "FFEBEE"
COR_AMAR_CLR  = "FFFDE7"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Arial")

def border_thin():
    s = Side(style="thin", color=COR_CINZA_BD)
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def header_cell(ws, row, col, value, bg=COR_ROXO_ESC, fg=COR_BRANCO, bold=True, size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg)
    c.font = font(bold=bold, color=fg, size=size)
    c.alignment = center()
    c.border = border_thin()
    return c

def data_cell(ws, row, col, value, bg=COR_BRANCO, fg="000000",
              bold=False, align="center", size=10, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg)
    c.font = Font(bold=bold, italic=italic, color=fg, size=size, name="Arial")
    c.alignment = center() if align == "center" else left()
    c.border = border_thin()
    return c

# ── Leitura do banco ───────────────────────────────────────
def load_data(db_path):
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    participantes = conn.execute(
        "SELECT id, nome, enviado_em FROM participantes ORDER BY enviado_em"
    ).fetchall()

    palpites = {}
    for p in participantes:
        rows = conn.execute(
            "SELECT jogo, gols_brasil, gols_adversario FROM palpites WHERE participante_id=?",
            (p["id"],)
        ).fetchall()
        palpites[p["id"]] = {r["jogo"]: (r["gols_brasil"], r["gols_adversario"]) for r in rows}

    classif_palpites = {}
    for p in participantes:
        row = conn.execute(
            "SELECT primeiro, segundo, terceiro, quarto FROM classificacao_palpites WHERE participante_id=?",
            (p["id"],)
        ).fetchone()
        if row:
            classif_palpites[p["id"]] = list(row)

    placares_reais = {
        r["jogo"]: {"brasil": r["gols_brasil"], "adversario": r["gols_adversario"], "encerrado": r["encerrado"]}
        for r in conn.execute("SELECT jogo, gols_brasil, gols_adversario, encerrado FROM placares_reais").fetchall()
    }

    classif_real = {
        r["posicao"]: r["time"]
        for r in conn.execute("SELECT posicao, time FROM classificacao_real ORDER BY posicao").fetchall()
    }

    conn.close()
    return participantes, palpites, classif_palpites, placares_reais, classif_real

# ── Pontuação (mesma lógica do scoring.py) ─────────────────
JOGOS_ADV = {
    "jogo1": "Marrocos",
    "jogo2": "Haiti",
    "jogo3": "Escócia",
}

def calcular_pontos(pid, palpites, classif_palpites, placares_reais, classif_real):
    total = 0
    detalhes_jogos = {}

    for jogo in ["jogo1", "jogo2", "jogo3"]:
        real = placares_reais.get(jogo, {})
        if not real.get("encerrado") or jogo not in palpites.get(pid, {}):
            detalhes_jogos[jogo] = {"pts": 0, "status": "pendente"}
            continue
        pb, pa = palpites[pid][jogo]
        rb, ra = real["brasil"], real["adversario"]
        if pb == rb and pa == ra:
            detalhes_jogos[jogo] = {"pts": 50, "status": "exato"}
            total += 50
        else:
            detalhes_jogos[jogo] = {"pts": 0, "status": "errado"}

    pts_classif = 0
    acertos_classif = {}
    gabarito = False
    if classif_real and len(classif_real) == 4 and pid in classif_palpites:
        pal = classif_palpites[pid]
        real_lista = [classif_real[i] for i in range(1, 5)]
        acertos = 0
        for i in range(4):
            if pal[i].upper() == real_lista[i].upper():
                acertos_classif[i+1] = True
                pts_classif += 30
                acertos += 1
            else:
                acertos_classif[i+1] = False
        if acertos == 4:
            pts_classif += 100
            gabarito = True
    total += pts_classif
    return total, detalhes_jogos, pts_classif, acertos_classif, gabarito

# ── Aba 1: Ranking ─────────────────────────────────────────
def aba_ranking(wb, participantes, palpites, classif_palpites, placares_reais, classif_real):
    ws = wb.create_sheet("🏆 Ranking")
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 22

    # Título
    ws.merge_cells("A1:L1")
    c = ws["A1"]
    c.value = "🇧🇷  BOLÃO COPA DO MUNDO 2026 — GRUPO C  |  RANKING GERAL"
    c.fill = fill(COR_ROXO_ESC)
    c.font = Font(bold=True, color=COR_AMARELO, size=14, name="Arial")
    c.alignment = center()

    # Cabeçalhos
    headers = [
        "Pos.", "Nome", "Total Pts",
        "Jogo 1\n(Marrocos)", "Jogo 2\n(Haiti)", "Jogo 3\n(Escócia)",
        "Pts Jogos", "Pts Classif.",
        "1º Acerto", "2º Acerto", "3º Acerto", "4º Acerto"
    ]
    for col, h in enumerate(headers, 1):
        header_cell(ws, 2, col, h)

    widths = [6, 24, 11, 14, 14, 14, 10, 12, 10, 10, 10, 10]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)

    # Calcula ranking
    scores = []
    for p in participantes:
        pid = p["id"]
        total, det_j, pts_c, acert_c, gab = calcular_pontos(
            pid, palpites, classif_palpites, placares_reais, classif_real)
        scores.append((p, total, det_j, pts_c, acert_c, gab))
    scores.sort(key=lambda x: (-x[1], x[0]["enviado_em"]))

    medal_cores = {1: COR_OURO, 2: COR_PRATA, 3: COR_BRONZE}

    for rank, (p, total, det_j, pts_c, acert_c, gab) in enumerate(scores, 1):
        row = rank + 2
        ws.row_dimensions[row].height = 20
        bg = COR_CINZA_CLR if rank % 2 == 0 else COR_BRANCO
        medal_bg = medal_cores.get(rank)

        # Posição
        c_pos = data_cell(ws, row, 1, rank, bg=medal_bg or bg,
                          fg=COR_BRANCO if medal_bg else "333333", bold=bool(medal_bg))

        # Nome
        data_cell(ws, row, 2, p["nome"], bg=bg, fg="111111", bold=(rank <= 3), align="left")

        # Total pts
        data_cell(ws, row, 3, total, bg=COR_AMAR_CLR if total > 0 else bg,
                  fg=COR_ROXO_ESC, bold=True, size=11)

        # Jogos
        pts_jogos = 0
        for j_idx, jogo in enumerate(["jogo1", "jogo2", "jogo3"], 4):
            d = det_j.get(jogo, {})
            st = d.get("status", "pendente")
            pts = d.get("pts", 0)
            pts_jogos += pts
            if st == "exato":
                pal = palpites.get(p["id"], {}).get(jogo, ("?", "?"))
                txt = f"✅ {pal[0]}×{pal[1]}  (+50)"
                data_cell(ws, row, j_idx, txt, bg=COR_VERDE_CLR, fg="1B5E20", bold=True)
            elif st == "errado":
                pal = palpites.get(p["id"], {}).get(jogo, ("?", "?"))
                real = placares_reais.get(jogo, {})
                txt = f"❌ {pal[0]}×{pal[1]}  (real: {real.get('brasil','?')}×{real.get('adversario','?')})"
                data_cell(ws, row, j_idx, txt, bg=COR_VERM_CLR, fg="B71C1C")
            else:
                pal = palpites.get(p["id"], {}).get(jogo, None)
                txt = f"⏳ {pal[0]}×{pal[1]}" if pal else "⏳ —"
                data_cell(ws, row, j_idx, txt, bg=bg, fg="888888", italic=True)

        data_cell(ws, row, 7, pts_jogos, bg=bg, fg="333333", bold=pts_jogos > 0)
        data_cell(ws, row, 8, pts_c,
                  bg=COR_VERDE_CLR if pts_c > 0 else bg,
                  fg="1B5E20" if pts_c > 0 else "888888", bold=pts_c > 0)

        # Acertos de classificação
        for pos_idx in range(1, 5):
            acertou = acert_c.get(pos_idx)
            if acertou is True:
                txt = "✅ +30"
                cell_bg, cell_fg = COR_VERDE_CLR, "1B5E20"
            elif acertou is False:
                txt = "❌"
                cell_bg, cell_fg = COR_VERM_CLR, "B71C1C"
            else:
                txt = "⏳"
                cell_bg, cell_fg = bg, "888888"
            data_cell(ws, row, 8 + pos_idx, txt, bg=cell_bg, fg=cell_fg, bold=(acertou is True))

    # Rodapé com data de geração
    last_row = len(scores) + 3
    ws.merge_cells(f"A{last_row}:L{last_row}")
    c = ws.cell(row=last_row, column=1,
                value=f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c.font = Font(italic=True, color="999999", size=9, name="Arial")
    c.alignment = left()

# ── Aba 2: Palpites Jogos ──────────────────────────────────
def aba_palpites_jogos(wb, participantes, palpites, placares_reais):
    ws = wb.create_sheet("⚽ Palpites Jogos")
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 22

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "PALPITES DE PLACAR — JOGOS DO BRASIL"
    c.fill = fill(COR_VERDE)
    c.font = Font(bold=True, color=COR_BRANCO, size=13, name="Arial")
    c.alignment = center()

    headers = ["Nome", "Enviado em",
               "Jogo 1 (13/06)\nBrasil × Marrocos", "Resultado J1",
               "Jogo 2 (19/06)\nBrasil × Haiti",    "Resultado J2",
               "Jogo 3 (24/06)\nBrasil × Escócia",  "Resultado J3"]
    for col, h in enumerate(headers, 1):
        header_cell(ws, 2, col, h, bg=COR_VERDE)

    for i, w in enumerate([22, 18, 20, 12, 20, 12, 20, 12], 1):
        set_col_width(ws, i, w)

    for idx, p in enumerate(participantes):
        row = idx + 3
        bg = COR_CINZA_CLR if idx % 2 == 0 else COR_BRANCO
        ws.row_dimensions[row].height = 18
        data_cell(ws, row, 1, p["nome"], bg=bg, fg="111111", align="left", bold=True)
        data_cell(ws, row, 2, p["enviado_em"], bg=bg, fg="444444", size=9)

        col = 3
        for jogo in ["jogo1", "jogo2", "jogo3"]:
            pal = palpites.get(p["id"], {}).get(jogo)
            real = placares_reais.get(jogo, {})
            if pal:
                txt_pal = f"Brasil {pal[0]} × {pal[1]}"
                if real.get("encerrado"):
                    acerto = pal[0] == real["brasil"] and pal[1] == real["adversario"]
                    txt_res = "✅ EXATO" if acerto else "❌ Errou"
                    bg_res = COR_VERDE_CLR if acerto else COR_VERM_CLR
                    fg_res = "1B5E20" if acerto else "B71C1C"
                else:
                    txt_res = "⏳ Aguardando"
                    bg_res, fg_res = bg, "888888"
            else:
                txt_pal = "—"
                txt_res = "—"
                bg_res, fg_res = bg, "AAAAAA"

            data_cell(ws, row, col,   txt_pal, bg=bg,    fg="222222")
            data_cell(ws, row, col+1, txt_res, bg=bg_res, fg=fg_res, bold=(txt_res == "✅ EXATO"))
            col += 2

# ── Aba 3: Palpites Classificação ─────────────────────────
def aba_palpites_classif(wb, participantes, classif_palpites, classif_real):
    ws = wb.create_sheet("🏅 Classificação")
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 22

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "PALPITES DE CLASSIFICAÇÃO — GRUPO C"
    c.fill = fill(COR_ROXO_MED)
    c.font = Font(bold=True, color=COR_AMARELO, size=13, name="Arial")
    c.alignment = center()

    headers = ["Nome", "1º Colocado", "Acerto 1º",
               "2º Colocado", "Acerto 2º",
               "3º Colocado", "Acerto 3º",
               "4º Colocado", "Acerto 4º", "Pts Classif."]
    for col, h in enumerate(headers, 1):
        header_cell(ws, 2, col, h, bg=COR_ROXO_MED)

    for i, w in enumerate([22, 14, 10, 14, 10, 14, 10, 14, 10, 12], 1):
        set_col_width(ws, i, w)

    real_lista = [classif_real.get(i, "—") for i in range(1, 5)] if len(classif_real) == 4 else []

    for idx, p in enumerate(participantes):
        row = idx + 3
        bg = COR_CINZA_CLR if idx % 2 == 0 else COR_BRANCO
        ws.row_dimensions[row].height = 18
        data_cell(ws, row, 1, p["nome"], bg=bg, fg="111111", align="left", bold=True)

        cl = classif_palpites.get(p["id"], [])
        pts_c = 0
        col = 2
        for i in range(4):
            time_pal = cl[i] if i < len(cl) else "—"
            data_cell(ws, row, col, time_pal, bg=bg, fg="222222")
            if real_lista and time_pal != "—":
                acerto = time_pal.upper() == real_lista[i].upper()
                if acerto:
                    pts_c += 30
                txt_a = "✅ +30" if acerto else "❌"
                bg_a  = COR_VERDE_CLR if acerto else COR_VERM_CLR
                fg_a  = "1B5E20"      if acerto else "B71C1C"
                data_cell(ws, row, col+1, txt_a, bg=bg_a, fg=fg_a, bold=acerto)
            else:
                data_cell(ws, row, col+1, "⏳", bg=bg, fg="888888")
            col += 2

        if pts_c == 120:  # 4×30 + bonus 100 nao está aqui mas pts_c só conta 30×n
            pts_c_txt = f"{pts_c} + 100 bônus"
        else:
            pts_c_txt = str(pts_c) if real_lista else "⏳"
        data_cell(ws, row, 10, pts_c_txt, bg=COR_AMAR_CLR if pts_c > 0 else bg,
                  fg=COR_ROXO_ESC, bold=pts_c > 0)

# ── Aba 4: Configurações ───────────────────────────────────
def aba_config(wb, placares_reais, classif_real):
    ws = wb.create_sheet("⚙️ Configurações")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "PLACARES REAIS E CLASSIFICAÇÃO FINAL REGISTRADOS"
    c.fill = fill(COR_ROXO_ESC)
    c.font = Font(bold=True, color=COR_AMARELO, size=13, name="Arial")
    c.alignment = center()

    # Placares
    header_cell(ws, 3, 1, "Jogo",       bg=COR_ROXO_MED)
    header_cell(ws, 3, 2, "Adversário", bg=COR_ROXO_MED)
    header_cell(ws, 3, 3, "Placar",     bg=COR_ROXO_MED)
    header_cell(ws, 3, 4, "Status",     bg=COR_ROXO_MED)

    for i, w in enumerate([12, 16, 20, 14], 1):
        set_col_width(ws, i, w)

    jogos_info = [
        ("jogo1", "Jogo 1 · 13/06", "Marrocos"),
        ("jogo2", "Jogo 2 · 19/06", "Haiti"),
        ("jogo3", "Jogo 3 · 24/06", "Escócia"),
    ]
    for row, (jkey, jlabel, jadv) in enumerate(jogos_info, 4):
        real = placares_reais.get(jkey, {})
        enc  = real.get("encerrado", 0)
        if enc:
            placar = f"Brasil {real['brasil']} × {real['adversario']}"
            status, bg_s, fg_s = "✅ Encerrado", COR_VERDE_CLR, "1B5E20"
        else:
            placar = "—"
            status, bg_s, fg_s = "⏳ Pendente", COR_AMAR_CLR, "795548"

        data_cell(ws, row, 1, jlabel, align="left")
        data_cell(ws, row, 2, jadv,   align="left")
        data_cell(ws, row, 3, placar, bold=enc)
        data_cell(ws, row, 4, status, bg=bg_s, fg=fg_s, bold=enc)

    # Classificação real
    ws.cell(row=8, column=1, value="").fill = fill(COR_BRANCO)
    header_cell(ws, 9, 1, "Posição",   bg=COR_VERDE)
    header_cell(ws, 9, 2, "Time",      bg=COR_VERDE)

    pos_cores = {1: COR_OURO, 2: COR_PRATA, 3: COR_BRONZE}
    pos_labels = {1: "🥇 1º Colocado", 2: "🥈 2º Colocado",
                  3: "🥉 3º Colocado", 4: "4️⃣ 4º Colocado"}
    for pos in range(1, 5):
        time = classif_real.get(pos, "—")
        bg_p = pos_cores.get(pos, COR_CINZA_CLR)
        data_cell(ws, 9+pos, 1, pos_labels[pos], bg=bg_p, fg="111111", bold=(pos <= 3))
        data_cell(ws, 9+pos, 2, time, bg=bg_p, fg="111111", bold=(pos <= 3), align="left")

# ── Main ────────────────────────────────────────────────────
def exportar(db_path, out_path):
    if not os.path.exists(db_path):
        print(f"❌ Banco '{db_path}' não encontrado.")
        sys.exit(1)

    print(f"📂 Lendo '{db_path}'...")
    participantes, palpites, classif_palpites, placares_reais, classif_real = load_data(db_path)

    print(f"   {len(participantes)} participante(s) encontrado(s).")

    wb = Workbook()
    wb.remove(wb.active)  # remove aba padrão

    aba_ranking(wb, participantes, palpites, classif_palpites, placares_reais, classif_real)
    aba_palpites_jogos(wb, participantes, palpites, placares_reais)
    aba_palpites_classif(wb, participantes, classif_palpites, classif_real)
    aba_config(wb, placares_reais, classif_real)

    wb.save(out_path)
    print(f"✅ Excel salvo em '{out_path}'")
    print(f"   Abas: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Exporta o bolão para Excel.")
    parser.add_argument("--db",  default="bolao.db",
                        help="Caminho do banco SQLite (padrão: bolao.db)")
    parser.add_argument("--out", default=None,
                        help="Arquivo de saída (padrão: bolao_YYYYMMDD_HHMM.xlsx)")
    args = parser.parse_args()

    if args.out is None:
        args.out = f"bolao_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    exportar(args.db, args.out)
